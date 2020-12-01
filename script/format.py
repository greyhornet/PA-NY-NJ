import os, time
from openpyxl import load_workbook, Workbook

path = "C:\\Users\\user\\Documents\\UiPath\\Ipass_Demo\\OutputFiles"
os.chdir(path)

def initialiseOPFile(justfilename):
    name = "..\\ProcessedOP\\"+justfilename+".xlsx"
    wb = Workbook()
    sheet = wb.active

    sheet["A1"] = "TOLL AGENCY"
    sheet["B1"] = "LP"
    sheet["C1"] = "LP STATE"
    sheet["D1"] = "TRXN. DATE & TIME"
    sheet["E1"] = "EXIT LANE"
    sheet["F1"] = "ACCOUNT #"
    sheet["G1"] = "REFERENCE # OR INVOICE #"
    sheet["H1"] = "VIOLATION #"
    sheet["I1"] = "AMOUNT DUE"
    sheet["J1"] = "DUE DATE"
    
    wb.save(filename=name)

    return name

def writerow(list_, OPfile):
    wb1 = load_workbook(OPfile)
    sheet1 = wb1.active
    end = str(int(sheet1.dimensions.split(":")[1][1:]) + 1)

    sheet1["A"+end] = list_[0]
    sheet1["B"+end] = list_[1]
    sheet1["C"+end] = list_[2]
    sheet1["D"+end] = list_[3]
    sheet1["E"+end] = list_[4]
    sheet1["F"+end] = list_[5]
    sheet1["G"+end] = list_[6]
    sheet1["H"+end] = list_[7]
    sheet1["I"+end] = list_[8]
    sheet1["J"+end] = list_[9]

    wb1.save(OPfile)

def justfilename(filename):
    return filename.split(".")[0]

def processFile(filename):
    file_=filename
    OP = initialiseOPFile(justfilename(file_))
    primary = load_workbook(filename=file_)
    sheet_p = primary.active
    rowsprocessed = 0

    for row in sheet_p.iter_rows(values_only=True):
        if row[1] != None:
            Agency = "ILLINOIS TOLLWAY"
            LP = row[7].split()[0]
            state = row[7].split()[1].replace("(","").replace(")","")
            datetime_ = row[1]+" "+row[2]
            exit_ = row[3]+" "+row[4]
            account = None
            invoice = row[6]
            violation = row[0]
            amount = row[5]
            duedate = row[8]

            array = [Agency, LP, state, datetime_, exit_, account, invoice, violation, amount, duedate]

            writerow(array, OP)
            rowsprocessed += 1
    #return rowsprocessed

#for i in os.listdir():
    #if i.endswith(".xlsx"):
        #start = time.time()
        #rows = processFile(i)
        #stop = time.time()
        #tt = stop-start
        #print(f'Done processing {rows} rows from {i} in {tt} seconds')
